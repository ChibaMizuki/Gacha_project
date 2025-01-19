import random, time, math
import openpyxl
import numpy as  np
from array import array


class Gacha():
    # 0 is lose, 1 is win
    def __init__(self, rate):
        self.rate = [rate, 100 - rate]
        self.gacha_list = [1, 0]
    
    # 仮想ガチャ、continuous 連をnumber_of_trials回行う
    # count how many times 1 appear
    def virtual_gacha(self, number_of_trials, continuous=1):
        num = 0
        for _ in range(number_of_trials):
            result = random.choices(self.gacha_list, k=continuous, weights=self.rate)
            num += result.count(1)
        
        return num
    
    # ガチャの確率を求める
    # calculate the probability
    def gacha_rate(self, number_of_trials):
        win = self.virtual_gacha(number_of_trials)
        rate = win / number_of_trials * 100
        
        return rate
    
    # 初めてあたりを引けるまで行うガチャ
    # Other experiment
    def first_get_count(self):
        get_count = 0
        result = 0
        
        while(result != 1):
            result = random.choices(self.gacha_list, weights=self.rate)[0]
            get_count += 1
        
        return get_count
    
    # 初めてあたりを引くことをgacha_cont 回行う
    # Other experiment
    def rate_first_get(self, number_of_trials):
        # 範囲の個数
        label_size = 15
        # 範囲
        range_size = 20
        
        ranges = array("i", [range_size * (x + 1) for x in range(label_size)])
        labels = [f"{range_size * x + 1} ~ {range_size * x + range_size}" for x in range(label_size)] + [f"{(label_size) * range_size + 1} ~ "]
        gacha_time = {label: 0 for label in labels}
        
        max = 0
        min = label_size * range_size
        average = 0
        count_list = array("i", [])
        print("\n初めて獲得するまで / Roll a gacha until getting item first")
        for i in range(number_of_trials):
            count = self.first_get_count()
            count_list.append(count)
            average += count
            if i % 100 == 0:
                print(f"\r{int(i / (number_of_trials / 100))}% completed", end="")
            # 最大値
            if count > max:
                max = count
            # 最小値
            if count < min:
                min = count
            # 辞書
            for i, limit in enumerate(ranges):
                if count <= limit:
                    gacha_time[labels[i]] += 1
                    break
            else:
                gacha_time[labels[-1]] += 1
        median = np.median(count_list)
        average /= number_of_trials
        print("\r100% completed      ")
        # 試行回数の辞書、最大試行回数、最小試行回数、中央値、平均値
        return gacha_time, max, min, median, average
    
    # 課金額以内で当たりを引ける確率
    # calculate the probability of winning within budget
    def budget_gacha(self, number_of_trials, number_of_gacha):
        get = 0
        get = array("i", [])
        print("\n課金 / Budget Gacha")
        # number_of_gacha 連をnumber_of_trials 回引く
        for i in range(number_of_trials):
            win = self.virtual_gacha(number_of_trials=1, continuous=number_of_gacha)
            get.append(win)
            print(f"\r{int(i  / (number_of_trials / 100))}% completed", end="")

        print("\r100% completed    \n")
        return get


# Function to export to Excel
def export_to_excel(result_list, file_name="gacha_results.xlsx"):
    try:
        # 既存のExcelファイルを開く
        wb = openpyxl.load_workbook(file_name)
        print(f"Existing file '{file_name}' loaded successfully.")
    except FileNotFoundError:
        # ファイルが存在しない場合は新しいファイルを作成
        wb = openpyxl.Workbook()
        print(f"File '{file_name}' not found. Creating a new file.")
    
    # シートをリセットまたは新規作成
    if "Gacha Results" in wb.sheetnames:
        ws = wb["Gacha Results"]
        ws.delete_rows(1, ws.max_row)  # 既存データを削除
        print(f"Sheet 'Gacha Results' reset.")
    else:
        ws = wb.create_sheet("Gacha Results")
        print(f"New sheet 'Gacha Results' created.")
    
    # データを書き込み
    ws.append(["Virtual Gacha Result"])
    ws.append(["確率", "", result_list[0]])
    ws.append(["平均試行回数", "", f"{result_list[5]}連"])
    ws.append(["中央値", "", f"{result_list[4]}連"])
    ws.append(["最小試行回数", "", f"{result_list[3]}連"])
    ws.append(["最大試行回数", "", f"{result_list[2]}連"])
    ws.append([])

    ws.append(["Budget Gacha Result"])
    ws.append(["獲得数", "獲得回数", "確率"])
    for i in range(len(result_list[8])):
        ws.append([f"{i}体", result_list[8][i]])
    ws.append([])

    ws.append(["Range", "Count times", "rate %", "comulative %"])
    i = 0
    for k, v in result_list[1].items():
        ws.append([k, v, result_list[6][i], result_list[7][i]])
        i += 1

    # ファイルを保存
    wb.save(file_name)
    print(f"Results saved to '{file_name}'.")



def main():
    # 試行回数 / Number of trials
    number_of_trials = int(input("試行回数 / Number of trials: "))
    # 確率 / Rate (rate %)
    rate = float(input("確率 / Rate (%): "))
    # 課金額 / Budget (budget yen)
    budget = int(input("課金額（円） / Budget (yen): "))
    # 1連の単価 / Cost of one roll 
    cost = 300
    # ガチャを回せる回数 / Number of time rolling gacha within budget
    number_of_gacha = int(budget / cost)
    
    gacha = Gacha(rate)

    start = time.time()

    rate_result = gacha.gacha_rate(number_of_trials)
    rate_result_dict, max, min, median, average_rate = gacha.rate_first_get(number_of_trials)
    budget_gacha_rate_list = gacha.budget_gacha(number_of_trials, number_of_gacha)

    end = time.time()
    
    print("-" * 50 + "Other experiment" + "-" * 50)
    print(f"{rate}% の当たり確率で最初にあたりを引くまでに回した回数 / Number of times until winning (gacha with a {rate}% chance of winning)")
    total_rate = 0
    cumulative_rate_list = array("f", [])
    each_rate_list = array("f", [])
    for k, v in rate_result_dict.items():
        each_rate = v / number_of_trials * 100
        total_rate += each_rate
        cumulative_rate_list.append(total_rate)
        each_rate_list.append(each_rate)
        print(f"{k:^10} 連 / Rolls: {v:>8} 回 / times: {each_rate:>6.2f} %: {total_rate:>6.2f} %")
    print(f"\n試行時の確率 / Rate                    : {rate_result:>6.2f} %")
    print(f"試行回数 / Number of trials            : {number_of_trials:>6} 回 / times")
    print(f"中央値 / Median                        : {median:>6} 回 / times")
    print(f"平均値 / Average                       : {average_rate:>6.2f} 回 / times")
    print(f"最大試行回数 / Maximum number of trials: {max:>6} 回 / times")
    print(f"最小試行回数 / Minimum number of trials: {min:>6} 回 / times\n")
    print("-" * 110)
    
    
    # Try the "number_of_trials" times whether the character is pulled or not within th ebudget
    i = 0
    total = 0
    total_theory = 0
    probability = 0
    get_list = []
    print(f"\n{budget}円課金した場合（{number_of_gacha}連×{number_of_trials}回) / When charging {budget}yen ({number_of_gacha} rolls * {number_of_trials} times)")
    print("獲得数:      獲得回数  :     確率  :  理論値 / Number of wins: times: Percentage: Theoretical rate: Theoretical value")
    while i < 3:
        get = budget_gacha_rate_list.count(i)
        get_list.append(get)
        get_rate = get / len(budget_gacha_rate_list) * 100
        theory = ((100 - rate) / 100) ** (number_of_gacha - i) * (rate / 100) ** i * math.comb(number_of_gacha, i) * 100
        probability += i * get
        print(f"{i:^6}: {get:>6} 回/times: {get_rate:>8.2f} %: {theory:>6.2f} %")
        total += get
        total_theory += theory
        i += 1
    get = len(budget_gacha_rate_list) - total
    get_rate = get / len(budget_gacha_rate_list) * 100
    get_list.append(get)
    print(f"  {i} ~ : {get:>6} 回/times: {get_rate:>8.2f} %: {100 - total_theory:>6.2f} %")
    
    probability += i * get
    print(f"\nProbability: {probability / (number_of_gacha * number_of_trials) * 100:.3}%\n")
        
    
    # 試行時の確率、結果の辞書、最大試行回数、最小、中央値、平均、各確率、累積確率
    to_excel_list = [
        rate_result,
        rate_result_dict,
        max,
        min,
        median,
        average_rate,
        each_rate_list,
        cumulative_rate_list,
        get_list
        ]
    
    # Excelに書き込み 
    # export_to_excel(to_excel_list)
    
    
    time_diff = end - start
    print(f"\n{time_diff:.4f} seconds")


if __name__ == "__main__":
    main()
